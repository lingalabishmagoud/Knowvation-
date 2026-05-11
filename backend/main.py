from fastapi import FastAPI, BackgroundTasks, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from typing import List, Optional
from pydantic import BaseModel
from database import db, DATABASE_ID, SOURCE_JOBS_COLLECTION, ANALYZED_JOBS_COLLECTION
from scraper import scrape_jobs
from ai_engine import extract_job_info, test_api_key
import uuid
import io
import csv
import json

app = FastAPI(title="Knowvation API")

def get_docs(res):
    docs = []
    if isinstance(res, dict):
        docs = res.get('documents', [])
    elif hasattr(res, 'documents'):
        docs = getattr(res, 'documents')
    
    final_docs = []
    for d in docs:
        if isinstance(d, dict):
            final_docs.append(d)
        elif hasattr(d, '__dict__'):
            final_docs.append(d.__dict__)
        else:
            final_docs.append(d)
    return final_docs

# ─── Models ──────────────────────────────────────────────────────────────────

class SourceJob(BaseModel):
    # Public job info (shown on mock board)
    title: str
    company: str
    location: str = ""
    description: str
    # HR / Recruiter intel (internal only — never shown on public board)
    hr_name: str = ""
    hr_designation: str = ""
    hr_email: str = ""
    hr_linkedin: str = ""
    hr_contact: str = ""

# ─── CORS ────────────────────────────────────────────────────────────────────

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Health ──────────────────────────────────────────────────────────────────

@app.get("/")
async def root():
    return {"message": "Knowvation API is online"}

# ─── Source Jobs (Admin CRUD) ─────────────────────────────────────────────────

@app.post("/admin/add-job")
async def add_source_job(job: SourceJob):
    try:
        doc_id = str(uuid.uuid4())
        db.create_document(
            database_id=DATABASE_ID,
            collection_id=SOURCE_JOBS_COLLECTION,
            document_id=doc_id,
            data=job.dict()
        )
        return {"status": "success", "id": doc_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/source-jobs")
async def get_source_jobs():
    try:
        response = db.list_documents(
            database_id=DATABASE_ID,
            collection_id=SOURCE_JOBS_COLLECTION
        )
        return get_docs(response)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/admin/source-jobs/{job_id}")
async def delete_source_job(job_id: str):
    try:
        db.delete_document(
            database_id=DATABASE_ID,
            collection_id=SOURCE_JOBS_COLLECTION,
            document_id=job_id
        )
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/admin/source-jobs/{job_id}")
async def update_source_job(job_id: str, job: SourceJob):
    try:
        db.update_document(
            database_id=DATABASE_ID,
            collection_id=SOURCE_JOBS_COLLECTION,
            document_id=job_id,
            data=job.dict()
        )
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ─── Recruiters Directory ────────────────────────────────────────────────────

@app.get("/admin/recruiters")
async def get_recruiters():
    """Return HR contact info extracted from source jobs — internal admin use."""
    try:
        response = db.list_documents(
            database_id=DATABASE_ID,
            collection_id=SOURCE_JOBS_COLLECTION
        )
        recruiters = []
        for doc in get_docs(response):
            if doc.get("hr_name") or doc.get("hr_email"):
                recruiters.append({
                    "id": doc["$id"],
                    "company": doc.get("company", ""),
                    "job_title": doc.get("title", ""),
                    "location": doc.get("location", ""),
                    "hr_name": doc.get("hr_name", ""),
                    "hr_designation": doc.get("hr_designation", ""),
                    "hr_email": doc.get("hr_email", ""),
                    "hr_linkedin": doc.get("hr_linkedin", ""),
                    "hr_contact": doc.get("hr_contact", ""),
                })
        return recruiters
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ─── Scraper ──────────────────────────────────────────────────────────────────

@app.post("/scrape")
def trigger_scrape():
    import subprocess
    import os
    try:
        subprocess.run(
            ["python", "test_all_scraper.py"],
            cwd=os.path.dirname(os.path.abspath(__file__)),
            check=True
        )
        return {"message": "Scraping completed successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/scrape-status")
async def get_scrape_status():
    import os
    status_file = os.path.join(os.path.dirname(__file__), "scrape_status.json")
    if os.path.exists(status_file):
        with open(status_file) as f:
            return json.load(f)
    return {"total": 0, "success": 0, "rate_limited": 0, "errors": 0, "message": "No scrape has been run yet.", "quota_warning": False}

# ─── AI Quota Status ──────────────────────────────────────────────────────────

@app.get("/api/quota-status")
async def get_quota_status():
    """Live-tests the current Gemini API key from .env and returns real status."""
    result = test_api_key()
    return result

# ─── Analytics ───────────────────────────────────────────────────────────────

@app.get("/analytics/jobs")
async def get_analyzed_jobs():
    try:
        response = db.list_documents(
            database_id=DATABASE_ID,
            collection_id=ANALYZED_JOBS_COLLECTION
        )
        return get_docs(response)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ─── Reports & Export ────────────────────────────────────────────────────────

@app.get("/reports/summary")
async def get_reports_summary():
    """Returns aggregated analytics data for the reports dashboard."""
    try:
        jobs_resp = db.list_documents(DATABASE_ID, ANALYZED_JOBS_COLLECTION)
        jobs = get_docs(jobs_resp)

        # Skill demand
        skill_counts: dict = {}
        for j in jobs:
            for s in (j.get("skills") or []):
                skill_counts[s] = skill_counts.get(s, 0) + 1

        # Location demand
        location_counts: dict = {}
        for j in jobs:
            loc = (j.get("location") or "Unknown").strip()
            if loc:
                location_counts[loc] = location_counts.get(loc, 0) + 1

        # Company hiring frequency
        company_counts: dict = {}
        for j in jobs:
            c = (j.get("company") or "Unknown").strip()
            if c and c != "—":
                company_counts[c] = company_counts.get(c, 0) + 1

        # Trend distribution
        trend_counts: dict = {}
        for j in jobs:
            t = j.get("hiring_trend") or "Unknown"
            trend_counts[t] = trend_counts.get(t, 0) + 1

        return {
            "total_jobs": len(jobs),
            "skill_demand": sorted(skill_counts.items(), key=lambda x: -x[1])[:15],
            "location_demand": sorted(location_counts.items(), key=lambda x: -x[1]),
            "company_frequency": sorted(company_counts.items(), key=lambda x: -x[1]),
            "trend_distribution": trend_counts,
            "avg_score": round(
                sum(j.get("intelligence_score", 0) for j in jobs) / len(jobs), 1
            ) if jobs else 0,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/export/csv")
async def export_csv():
    """Export all analyzed jobs as a CSV file download."""
    try:
        jobs_resp = db.list_documents(DATABASE_ID, ANALYZED_JOBS_COLLECTION)
        jobs = get_docs(jobs_resp)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Role", "Company", "Location", "Experience", "Hiring Trend", "Intelligence Score", "Skills", "HR Name", "HR Email"])
        for j in jobs:
            writer.writerow([
                j.get("role", ""),
                j.get("company", ""),
                j.get("location", ""),
                j.get("experience", ""),
                j.get("hiring_trend", ""),
                j.get("intelligence_score", 0),
                ", ".join(j.get("skills") or []),
                j.get("hr_name", ""),
                j.get("hr_email", ""),
            ])

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=knowvation_jobs_report.csv"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/export/excel")
async def export_excel():
    """Export all analyzed jobs as an Excel file download."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        jobs_resp = db.list_documents(DATABASE_ID, ANALYZED_JOBS_COLLECTION)
        jobs = get_docs(jobs_resp)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hiring Intelligence Report"

        headers = ["Role", "Company", "Location", "Experience", "Hiring Trend", "Score", "Skills", "HR Name", "HR Email", "HR Contact"]
        header_fill = PatternFill("solid", fgColor="6D28D9")
        header_font = Font(color="FFFFFF", bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = 20

        for row_idx, j in enumerate(jobs, 2):
            ws.cell(row=row_idx, column=1, value=j.get("role", ""))
            ws.cell(row=row_idx, column=2, value=j.get("company", ""))
            ws.cell(row=row_idx, column=3, value=j.get("location", ""))
            ws.cell(row=row_idx, column=4, value=j.get("experience", ""))
            ws.cell(row=row_idx, column=5, value=j.get("hiring_trend", ""))
            ws.cell(row=row_idx, column=6, value=j.get("intelligence_score", 0))
            ws.cell(row=row_idx, column=7, value=", ".join(j.get("skills") or []))
            ws.cell(row=row_idx, column=8, value=j.get("hr_name", ""))
            ws.cell(row=row_idx, column=9, value=j.get("hr_email", ""))
            ws.cell(row=row_idx, column=10, value=j.get("hr_contact", ""))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=knowvation_report.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
